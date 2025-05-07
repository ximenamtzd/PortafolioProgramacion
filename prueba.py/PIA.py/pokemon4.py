import tkinter as tk
import requests
import json
import csv
import re
import os
import statistics
import numpy as np
import matplotlib.pyplot as plt
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

# VALIDACIONES
regex_nombre = re.compile(r"^[a-zA-Z\-]+$")
regex_tipo = re.compile(r"^[a-z]+$")
regex_habilidad = re.compile(r"^[a-z\-]+$")

datos_pokemones = []
archivo_json = "pokemones.json"
archivo_csv = "pokemones.csv"
archivo_excel = "pokemones.xlsx"

#FUNCIONES
def validar_nombre(nombre):
    return bool(regex_nombre.fullmatch(nombre))

def obtener_datos_pokemon(nombre_pokemon):
    url = f"https://pokeapi.co/api/v2/pokemon/{nombre_pokemon.lower()}"
    respuesta = requests.get(url)
    if respuesta.status_code == 200:
        datos = respuesta.json()
        return {
            "nombre": datos["name"],
            "tipos": [t["type"]["name"] for t in datos["types"]],
            "habilidades": [h["ability"]["name"] for h in datos["abilities"]],
        }
    return None

def mostrar_info():
    nombre = entrada.get().strip()
    if not validar_nombre(nombre):
        messagebox.showerror("Error", "Nombre inválido. Usa solo letras sin espacios.")
        return

    datos = obtener_datos_pokemon(nombre)
    if datos:
        datos_pokemones.append(datos)
        texto = (
            f"Nombre: {datos['nombre'].capitalize()}\n"
            f"Tipos: {', '.join(datos['tipos'])}\n"
            f"Habilidades: {', '.join(datos['habilidades'])}"
        )
        resultado.config(text=texto)
        entrada.delete(0, tk.END)
    else:
        resultado.config(text=f"No se encontró información para: {nombre}")

def guardar_archivos():
    if not datos_pokemones:
        messagebox.showwarning("Aviso", "No hay datos para guardar.")
        return

    with open(archivo_json, "w") as f_json:
        json.dump(datos_pokemones, f_json, indent=4)

    with open(archivo_csv, "w", newline="") as f_csv:
        writer = csv.writer(f_csv)
        writer.writerow(["nombre", "tipos", "habilidades"])
        for p in datos_pokemones:
            writer.writerow([p["nombre"], ', '.join(p["tipos"]), ', '.join(p["habilidades"])])

    if os.path.exists(archivo_excel):
        wb = load_workbook(archivo_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pokemones"
        ws.append(["Nombre", "Tipos", "Habilidades"])

    for p in datos_pokemones:
        ws.append([p["nombre"], ', '.join(p["tipos"]), ', '.join(p["habilidades"])])

    wb.save(archivo_excel)

    messagebox.showinfo("Guardado", "Datos guardados exitosamente.")
    try:
        os.startfile(archivo_excel)
    except Exception as e:
        print(f"No se pudo abrir automáticamente el archivo: {e}")

    analizar_y_graficar()

def leer_desde_excel():
    if not os.path.exists(archivo_excel):
        return []

    wb = load_workbook(archivo_excel)
    ws = wb.active
    pokemones = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre, tipos, habilidades = fila
        if not (nombre and tipos and habilidades):
            continue
        pokemones.append({
            "nombre": nombre,
            "tipos": [t.strip() for t in tipos.split(',')],
            "habilidades": [h.strip() for h in habilidades.split(',')]
        })
    return pokemones

def validar_datos(pokemones):
    validos = []
    for p in pokemones:
        if (
            regex_nombre.fullmatch(p["nombre"]) and
            all(regex_tipo.fullmatch(t) for t in p["tipos"]) and
            all(regex_habilidad.fullmatch(h) for h in p["habilidades"])
        ):
            validos.append(p)
    return validos

def analizar_y_graficar():
    pokemones = leer_desde_excel()
    pokemones_validos = validar_datos(pokemones)
    if not pokemones_validos:
        print("No hay datos válidos.")
        return

    nombres = [p["nombre"].capitalize() for p in pokemones_validos]
    tipos_count = [len(p["tipos"]) for p in pokemones_validos]
    habilidades_count = [len(p["habilidades"]) for p in pokemones_validos]

    # Gráfica de líneas (tipos por Pokémon)
    plt.figure(figsize=(8, 5))
    plt.plot(nombres, tipos_count, marker='o', linestyle='-', color='blue')
    plt.title("Cantidad de Tipos por Pokémon")
    plt.xlabel("Pokémon")
    plt.ylabel("Número de Tipos")
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    # Gráfica de barras (habilidades por Pokémon)
    plt.figure(figsize=(8, 5))
    plt.bar(nombres, habilidades_count, color='green')
    plt.title("Cantidad de Habilidades por Pokémon")
    plt.xlabel("Pokémon")
    plt.ylabel("Número de Habilidades")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    # Diagrama de dispersión
    plt.figure(figsize=(8, 5))
    plt.scatter(tipos_count, habilidades_count, color='purple')
    plt.title("Dispersión de Tipos vs Habilidades")
    plt.xlabel("Número de Tipos")
    plt.ylabel("Número de Habilidades")
    plt.grid(True)
    plt.tight_layout()
    plt.show()

    # Gráfico de pastel (proporción de pokémon por cantidad de tipos)
    plt.figure(figsize=(6, 6))
    etiquetas = list(set(tipos_count))
    valores = [tipos_count.count(e) for e in etiquetas]
    plt.pie(valores, labels=[f"{e} tipo(s)" for e in etiquetas], autopct='%1.1f%%')
    plt.title("Distribución de Pokémon por Cantidad de Tipos")
    plt.tight_layout()
    plt.show()

# INTERFAZ GRÁFICA 
ventana = tk.Tk()
ventana.title("Consulta Pokémon")

tk.Label(ventana, text="Nombre del Pokémon:").pack(pady=5)
entrada = tk.Entry(ventana)
entrada.pack(pady=5)

tk.Button(ventana, text="Buscar", command=mostrar_info).pack(pady=5)
tk.Button(ventana, text="Guardar Datos y Analizar", command=guardar_archivos).pack(pady=5)

resultado = tk.Label(ventana, text="", justify="left")
resultado.pack(pady=10)

ventana.mainloop()
