import tkinter as tk
import requests
import json
import csv
import os
import re
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

# Carpeta donde se guardan los resultados
CARPETA_RESULTADOS = os.path.join(os.getcwd(), "resultados")
ARCHIVO_EXCEL = os.path.join(CARPETA_RESULTADOS, "pokemones.xlsx")
ARCHIVO_JSON = os.path.join(CARPETA_RESULTADOS, "pokemones.json")
ARCHIVO_CSV = os.path.join(CARPETA_RESULTADOS, "pokemones.csv")

# Asegurar existencia de la carpeta
if not os.path.exists(CARPETA_RESULTADOS):
    os.makedirs(CARPETA_RESULTADOS)

# Lista en memoria
datos_pokemones = []

# Cargar datos desde Excel si existe
def cargar_excel():
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            datos_pokemones.append({
                "nombre": row[0],
                "tipos": row[1].split(", "),
                "habilidades": row[2].split(", "),
                "altura": row[3],
                "peso": row[4],
                "experiencia_base": row[5]
            })

# Validar nombres sin caracteres raros
def validar_nombre(nombre):
    return bool(re.fullmatch(r"[a-zA-Z\-]+", nombre))

# Obtener información desde la API
def obtener_datos_pokemon(nombre_pokemon):
    url = f"https://pokeapi.co/api/v2/pokemon/{nombre_pokemon.lower()}"
    try:
        respuesta = requests.get(url, timeout=10)
        if respuesta.status_code == 200:
            datos = respuesta.json()
            return {
                "nombre": datos["name"],
                "tipos": [t["type"]["name"] for t in datos["types"]],
                "habilidades": [h["ability"]["name"] for h in datos["abilities"]],
                "altura": datos.get("height", 0),
                "peso": datos.get("weight", 0),
                "experiencia_base": datos.get("base_experience", 0)
            }
        elif respuesta.status_code == 404:
            return None
        else:
            raise requests.RequestException("Respuesta inesperada del servidor")
    except requests.exceptions.RequestException as e:
        raise

# Mostrar en pantalla
def mostrar_info():
    nombre = entrada.get().strip().lower()

    if not nombre:
        messagebox.showerror("Campo vacío", "Por favor, escribe el nombre de un Pokémon.")
        return

    if not validar_nombre(nombre):
        messagebox.showerror("Nombre inválido", "El nombre solo puede contener letras (sin espacios, números ni símbolos).")
        return

    if any(p["nombre"] == nombre for p in datos_pokemones):
        messagebox.showinfo("Repetido", f"{nombre.capitalize()} ya está registrado en la base de datos.")
        return

    try:
        datos = obtener_datos_pokemon(nombre)
    except requests.exceptions.RequestException:
        messagebox.showerror("Error de conexión", "No se pudo conectar a la API de Pokémon. Verifica tu conexión a internet.")
        return

    if datos:
        datos_pokemones.append(datos)
        texto = (
            f"Nombre: {datos['nombre'].capitalize()}\n"
            f"Tipos: {', '.join(datos['tipos'])}\n"
            f"Habilidades: {', '.join(datos['habilidades'])}\n"
            f"Altura: {datos['altura']}\n"
            f"Peso: {datos['peso']}\n"
            f"Experiencia base: {datos['experiencia_base']}"
        )
        resultado.config(text=texto)
        entrada.delete(0, tk.END)
    else:
        messagebox.showwarning("No encontrado", f"No se encontró información del Pokémon '{nombre}'. Verifica que esté bien escrito.")
        resultado.config(text="")

# Guardar en JSON, CSV y Excel (evitando duplicados)
def guardar_archivos():
    if not datos_pokemones:
        messagebox.showwarning("Aviso", "No hay datos para guardar.")
        return

    nombres_existentes = set()

    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombres_existentes.add(row[0])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pokemones"
        ws.append(["Nombre", "Tipos", "Habilidades", "Altura", "Peso", "Experiencia Base"])

    nuevos = 0
    for p in datos_pokemones:
        if p["nombre"] not in nombres_existentes:
            ws.append([
                p["nombre"],
                ', '.join(p["tipos"]),
                ', '.join(p["habilidades"]),
                p["altura"],
                p["peso"],
                p["experiencia_base"]
            ])
            nombres_existentes.add(p["nombre"])
            nuevos += 1

    wb.save(ARCHIVO_EXCEL)

    # JSON
    with open(ARCHIVO_JSON, "w") as f_json:
        json.dump(datos_pokemones, f_json, indent=4)

    # CSV
    with open(ARCHIVO_CSV, "w", newline="") as f_csv:
        writer = csv.writer(f_csv)
        writer.writerow(["nombre", "tipos", "habilidades", "altura", "peso", "experiencia_base"])
        for p in datos_pokemones:
            writer.writerow([
                p["nombre"],
                ', '.join(p["tipos"]),
                ', '.join(p["habilidades"]),
                p["altura"],
                p["peso"],
                p["experiencia_base"]
            ])

    messagebox.showinfo("Guardado", f"Datos guardados. {nuevos} nuevos añadidos.")

# Abrir la carpeta de resultados
def abrir_carpeta():
    try:
        os.startfile(CARPETA_RESULTADOS)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir la carpeta de resultados. Intenta hacerlo manualmente.\n\n{e}")

# Eliminar todo lo guardado
def eliminar_datos():
    errores = []
    for archivo in [ARCHIVO_EXCEL, ARCHIVO_JSON, ARCHIVO_CSV]:
        try:
            if os.path.exists(archivo):
                os.remove(archivo)
        except Exception as e:
            errores.append(str(e))

    datos_pokemones.clear()

    if errores:
        messagebox.showwarning("Parcialmente eliminado", f"Algunos archivos no pudieron eliminarse:\n\n{chr(10).join(errores)}")
    else:
        messagebox.showinfo("Eliminado", "Todos los datos guardados han sido eliminados exitosamente.")

# Leer datos desde Excel para graficar
def leer_datos_para_graficas():
    if not os.path.exists(ARCHIVO_EXCEL):
        return []

    datos = []
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        datos.append({
            "nombre": row[0].capitalize(),
            "altura": row[3],
            "peso": row[4],
            "experiencia_base": row[5]
        })
    return datos

# Mostrar gráficas desde Excel
def mostrar_graficas():
    datos = leer_datos_para_graficas()
    if not datos:
        messagebox.showwarning("Aviso", "No hay datos en Excel para graficar.")
        return

    nombres = [d["nombre"] for d in datos]
    alturas = [d["altura"] for d in datos]
    pesos = [d["peso"] for d in datos]
    experiencia = [d["experiencia_base"] for d in datos]

    plt.figure(figsize=(10, 6))
    plt.bar(nombres, alturas, color='skyblue')
    plt.title("Altura de los Pokémon")
    plt.xlabel("Pokémon")
    plt.ylabel("Altura")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    plt.figure(figsize=(10, 6))
    plt.bar(nombres, pesos, color='lightgreen')
    plt.title("Peso de los Pokémon")
    plt.xlabel("Pokémon")
    plt.ylabel("Peso")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    plt.figure(figsize=(10, 6))
    plt.bar(nombres, experiencia, color='salmon')
    plt.title("Experiencia base de los Pokémon")
    plt.xlabel("Pokémon")
    plt.ylabel("Experiencia Base")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

# GUI 

ventana = tk.Tk()
ventana.title("Consulta Pokémon")

tk.Label(ventana, text="Nombre del Pokémon:").pack(pady=5)
entrada = tk.Entry(ventana)
entrada.pack(pady=5)

tk.Button(ventana, text="Buscar Pokémon", command=mostrar_info).pack(pady=5)
tk.Button(ventana, text="Guardar Datos", command=guardar_archivos).pack(pady=5)
tk.Button(ventana, text="Mostrar Gráficas", command=mostrar_graficas).pack(pady=5)
tk.Button(ventana, text="Abrir Carpeta Resultados", command=abrir_carpeta).pack(pady=5)
tk.Button(ventana, text="Eliminar datos guardados", command=eliminar_datos).pack(pady=5)

resultado = tk.Label(ventana, text="", justify="left")
resultado.pack(pady=10)

cargar_excel()
ventana.mainloop()
