import tkinter as tk
import requests
import json
import csv
import re
import os
from tkinter import messagebox

# Lista para almacenar múltiples pokémon
datos_pokemones = []

# Validación con expresión regular (solo letras, sin espacios ni números)
def validar_nombre(nombre):
    return bool(re.fullmatch(r"[a-zA-Z\-]+", nombre))

# Función para obtener y procesar datos de la API
def obtener_datos_pokemon(nombre_pokemon):
    url = f"https://pokeapi.co/api/v2/pokemon/{nombre_pokemon.lower()}"
    respuesta = requests.get(url)

    if respuesta.status_code == 200:
        datos = respuesta.json()
        pokemon = {
            "nombre": datos["name"],
            "tipos": [t["type"]["name"] for t in datos["types"]],
            "habilidades": [h["ability"]["name"] for h in datos["abilities"]],
        }
        return pokemon
    else:
        return None

# Función para mostrar y guardar la información
def mostrar_info():
    nombre = entrada.get().strip()

    if not validar_nombre(nombre):
        messagebox.showerror("Error", "Nombre inválido. Usa solo letras sin espacios.")
        return

    datos = obtener_datos_pokemon(nombre)

    if datos:
        datos_pokemones.append(datos)
        texto = (
            f"Nombre: {datos['nombre'].capitalize()}\n"
            f"Tipos: {', '.join(datos['tipos'])}\n"
            f"Habilidades: {', '.join(datos['habilidades'])}"
        )
        resultado.config(text=texto)
        entrada.delete(0, tk.END)
    else:
        resultado.config(text=f"No se encontró información para: {nombre}")

# Función para guardar en archivos
from openpyxl import Workbook, load_workbook
import os

def guardar_archivos():
    if not datos_pokemones:
        messagebox.showwarning("Aviso", "No hay datos para guardar.")
        return

    # Guardar JSON
    with open("pokemones.json", "w") as f_json:
        json.dump(datos_pokemones, f_json, indent=4)

    # Guardar CSV plano
    with open("pokemones.csv", "w", newline="") as f_csv:
        writer = csv.writer(f_csv)
        writer.writerow(["nombre", "tipos", "habilidades"])
        for p in datos_pokemones:
            writer.writerow([p["nombre"], ', '.join(p["tipos"]), ', '.join(p["habilidades"])])

    # Guardar con openpyxl en Excel (no sobreescribir)
    ruta_excel = "pokemones.xlsx"
    if os.path.exists(ruta_excel):
        wb = load_workbook(ruta_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pokemones"
        ws.append(["Nombre", "Tipos", "Habilidades"])  # encabezado solo si es nuevo

    # Agregar nuevas filas al Excel
    for p in datos_pokemones:
        ws.append([p["nombre"], ', '.join(p["tipos"]), ', '.join(p["habilidades"])])

    wb.save(ruta_excel)

    # Abrir el archivo
    try:
        os.startfile(ruta_excel)
    except Exception as e:
        print(f"No se pudo abrir automáticamente el archivo: {e}")

    messagebox.showinfo("Guardado", "Datos guardados en JSON, CSV y Excel (.xlsx)")

# Interfaz gráfica
ventana = tk.Tk()
ventana.title("Consulta Pokémon")

tk.Label(ventana, text="Nombre del Pokémon:").pack(pady=5)
entrada = tk.Entry(ventana)
entrada.pack(pady=5)

tk.Button(ventana, text="Buscar", command=mostrar_info).pack(pady=5)
tk.Button(ventana, text="Guardar Datos", command=guardar_archivos).pack(pady=5)

resultado = tk.Label(ventana, text="", justify="left")
resultado.pack(pady=10)

ventana.mainloop()
