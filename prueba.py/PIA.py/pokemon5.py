import tkinter as tk
import requests
import json
import csv
import os
import re
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from collections import Counter

# Carpeta donde se guardan los resultados
CARPETA_RESULTADOS = os.path.join(os.getcwd(), "resultados")
ARCHIVO_EXCEL = os.path.join(CARPETA_RESULTADOS, "pokemones.xlsx")
ARCHIVO_JSON = os.path.join(CARPETA_RESULTADOS, "pokemones.json")
ARCHIVO_CSV = os.path.join(CARPETA_RESULTADOS, "pokemones.csv")

if not os.path.exists(CARPETA_RESULTADOS):
    os.makedirs(CARPETA_RESULTADOS)

datos_pokemones = []

def cargar_excel():
    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            datos_pokemones.append({
                "nombre": row[0],
                "tipos": row[1].split(", "),
                "habilidades": row[2].split(", "),
                "altura": row[3],
                "peso": row[4],
                "experiencia_base": row[5]
            })

def validar_nombre(nombre):
    return bool(re.fullmatch(r"[a-zA-Z\-]+", nombre))

def obtener_datos_pokemon(nombre_pokemon):
    url = f"https://pokeapi.co/api/v2/pokemon/{nombre_pokemon.lower()}"
    try:
        respuesta = requests.get(url)
        if respuesta.status_code == 200:
            datos = respuesta.json()
            return {
                "nombre": datos["name"],
                "tipos": [t["type"]["name"] for t in datos["types"]],
                "habilidades": [h["ability"]["name"] for h in datos["abilities"]],
                "altura": datos.get("height", 0),
                "peso": datos.get("weight", 0),
                "experiencia_base": datos.get("base_experience", 0)
            }
        else:
            return None
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Error", f"No se pudo conectar con la API:\n{e}")
        return None

def mostrar_info():
    nombre = entrada.get().strip().lower()
    if not validar_nombre(nombre):
        messagebox.showerror("Error", "Nombre inválido. Usa solo letras sin espacios.")
        return

    if any(p["nombre"] == nombre for p in datos_pokemones):
        messagebox.showinfo("Repetido", f"{nombre.capitalize()} ya está en la base de datos.")
        return

    datos = obtener_datos_pokemon(nombre)
    if datos:
        datos_pokemones.append(datos)
        texto = (
            f"Nombre: {datos['nombre'].capitalize()}\n"
            f"Tipos: {', '.join(datos['tipos'])}\n"
            f"Habilidades: {', '.join(datos['habilidades'])}\n"
            f"Altura: {datos['altura']}\n"
            f"Peso: {datos['peso']}\n"
            f"Experiencia base: {datos['experiencia_base']}"
        )
        resultado.config(text=texto)
        entrada.delete(0, tk.END)
    else:
        messagebox.showwarning("No encontrado", f"No se encontró información para: {nombre}")

def guardar_archivos():
    if not datos_pokemones:
        messagebox.showwarning("Aviso", "No hay datos para guardar.")
        return

    nombres_existentes = set()

    if os.path.exists(ARCHIVO_EXCEL):
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombres_existentes.add(row[0])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pokemones"
        ws.append(["Nombre", "Tipos", "Habilidades", "Altura", "Peso", "Experiencia Base"])

    nuevos = 0
    for p in datos_pokemones:
        if p["nombre"] not in nombres_existentes:
            ws.append([
                p["nombre"],
                ', '.join(p["tipos"]),
                ', '.join(p["habilidades"]),
                p["altura"],
                p["peso"],
                p["experiencia_base"]
            ])
            nombres_existentes.add(p["nombre"])
            nuevos += 1

    wb.save(ARCHIVO_EXCEL)

    with open(ARCHIVO_JSON, "w") as f_json:
        json.dump(datos_pokemones, f_json, indent=4)

    with open(ARCHIVO_CSV, "w", newline="") as f_csv:
        writer = csv.writer(f_csv)
        writer.writerow(["nombre", "tipos", "habilidades", "altura", "peso", "experiencia_base"])
        for p in datos_pokemones:
            writer.writerow([
                p["nombre"],
                ', '.join(p["tipos"]),
                ', '.join(p["habilidades"]),
                p["altura"],
                p["peso"],
                p["experiencia_base"]
            ])

    messagebox.showinfo("Guardado", f"Datos guardados. {nuevos} nuevos añadidos.")

def abrir_carpeta():
    try:
        os.startfile(CARPETA_RESULTADOS)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir la carpeta: {e}")

def eliminar_datos():
    for archivo in [ARCHIVO_EXCEL, ARCHIVO_JSON, ARCHIVO_CSV]:
        if os.path.exists(archivo):
            os.remove(archivo)
    datos_pokemones.clear()
    messagebox.showinfo("Eliminado", "Todos los datos guardados han sido eliminados.")

def leer_datos_para_graficas():
    if not os.path.exists(ARCHIVO_EXCEL):
        return []

    datos = []
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        datos.append({
            "nombre": row[0].capitalize(),
            "tipos": row[1].split(", "),
            "altura": row[3],
            "peso": row[4],
            "experiencia_base": row[5]
        })
    return datos

def mostrar_graficas():
    datos = leer_datos_para_graficas()
    if not datos:
        messagebox.showwarning("Aviso", "No hay datos en Excel para graficar.")
        return

    nombres = [d["nombre"] for d in datos]
    alturas = [d["altura"] for d in datos]
    pesos = [d["peso"] for d in datos]
    experiencia = [d["experiencia_base"] for d in datos]

    # 1. Gráfico de líneas: Experiencia base
    plt.figure(figsize=(10, 6))
    plt.plot(nombres, experiencia, marker='o', linestyle='-', color='purple')
    plt.title("Experiencia Base de los Pokémon")
    plt.xlabel("Pokémon")
    plt.ylabel("Experiencia Base")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    # 2. Gráfico de barras: Altura
    plt.figure(figsize=(10, 6))
    plt.bar(nombres, alturas, color='skyblue')
    plt.title("Altura de los Pokémon")
    plt.xlabel("Pokémon")
    plt.ylabel("Altura")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

    # 3. Diagrama de dispersión: Peso vs Altura
    plt.figure(figsize=(8, 6))
    plt.scatter(alturas, pesos, color='orange')
    plt.title("Relación entre Altura y Peso")
    plt.xlabel("Altura")
    plt.ylabel("Peso")
    plt.tight_layout()
    plt.show()

    # 4. Gráfico de pastel: Tipos
    todos_los_tipos = [tipo for d in datos for tipo in d["tipos"]]
    conteo_tipos = Counter(todos_los_tipos)
    plt.figure(figsize=(8, 8))
    plt.pie(conteo_tipos.values(), labels=conteo_tipos.keys(), autopct='%1.1f%%', startangle=90)
    plt.title("Distribución de Tipos de Pokémon")
    plt.tight_layout()
    plt.show()

# Interfaz gráfica
ventana = tk.Tk()
ventana.title("Consulta Pokémon")

tk.Label(ventana, text="Nombre del Pokémon:").pack(pady=5)
entrada = tk.Entry(ventana)
entrada.pack(pady=5)

tk.Button(ventana, text="Buscar Pokémon", command=mostrar_info).pack(pady=5)
tk.Button(ventana, text="Guardar Datos", command=guardar_archivos).pack(pady=5)
tk.Button(ventana, text="Mostrar Gráficas", command=mostrar_graficas).pack(pady=5)
tk.Button(ventana, text="Abrir Carpeta Resultados", command=abrir_carpeta).pack(pady=5)
tk.Button(ventana, text="Eliminar datos guardados", command=eliminar_datos).pack(pady=5)

resultado = tk.Label(ventana, text="", justify="left")
resultado.pack(pady=10)

cargar_excel()
ventana.mainloop()
